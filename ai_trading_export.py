"""
AI Trading experiment: Excel snapshot export + scoped trading reset.

Export does not modify any data.
Reset clears Paper Trading experiment state only — never AI Discovery / Saved News /
Watchlist / Research / scoring / system settings.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from db import get_conn, init_db

log = logging.getLogger("leibot.ai_trading_export")

PT = ZoneInfo("America/Los_Angeles")


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _pt_now_str() -> str:
    return datetime.now(PT).strftime("%Y-%m-%d %H:%M PT")


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([("" if v is None else v) for v in row])
    for i, h in enumerate(headers, start=1):
        width = max(12, min(42, len(str(h)) + 2))
        for r in rows[:80]:
            try:
                width = max(width, min(42, len(str(r[i - 1])) + 2))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(i)].width = width


def build_ai_trading_workbook() -> bytes:
    """Build .xlsx bytes for the current AI Trading experiment snapshot."""
    from ai_discovery import list_discovery_candidates
    from paper_trading import (
        format_source_label,
        holding_days_calendar,
        list_closed_trades,
        list_open_trades,
        normalize_exit_reason,
        portfolio_summary,
    )

    init_db()
    summary = portfolio_summary()
    opens = list_open_trades()
    closed = list_closed_trades(limit=5000)
    discovery = list_discovery_candidates(
        limit=500,
        recent_only=True,
        exclude_negative=False,
        history_mode=False,
    )
    # Saved News = ★ news priority (research pin), not Priority Buy.
    all_hist = list_discovery_candidates(
        limit=500,
        recent_only=False,
        exclude_negative=False,
        history_mode=True,
    )
    saved = [r for r in all_hist if int(r.get("is_news_priority") or 0)]

    wb = Workbook()
    # Remove default sheet; we create named sheets.
    default = wb.active
    wb.remove(default)

    exported_at = _pt_now_str()
    _write_sheet(
        wb,
        "Summary",
        ["Metric", "Value"],
        [
            ["Export date/time", exported_at],
            ["Starting Capital", summary.get("starting_capital")],
            ["Current Equity", summary.get("current_equity")],
            ["Cash", summary.get("cash")],
            ["Invested", summary.get("invested")],
            ["Total Realized P&L", summary.get("total_realized_pnl")],
            ["Total Unrealized P&L", summary.get("total_unrealized_pnl")],
            ["Total Return %", summary.get("total_return_pct")],
            ["Win Rate %", summary.get("win_rate")],
            ["Closed Trades", summary.get("closed_trades")],
            ["Open Positions", summary.get("open_trades")],
            ["Today's P&L", summary.get("today_pnl")],
        ],
    )

    hist_headers = [
        "Ticker",
        "Source",
        "Entry Date",
        "Entry Price",
        "Exit Date",
        "Exit Price",
        "Shares",
        "Cost",
        "P&L",
        "Return %",
        "Holding Days",
        "Exit Reason",
        "Stop Loss",
        "Take Profit",
        "AI Score at Entry",
        "Financial",
        "News",
        "Knife Risk",
        "63D at Entry",
        "Priority Buy",
    ]
    hist_rows: list[list[Any]] = []
    for t in closed:
        ok = t.get("financial_ok_entry")
        known = t.get("financial_known_entry")
        if ok is not None and known is not None:
            fin = f"{ok}/{known}"
        else:
            fin = t.get("financial_entry")
        hist_rows.append(
            [
                t.get("ticker"),
                format_source_label(t.get("source_at_entry")),
                t.get("entry_date"),
                t.get("entry_price"),
                t.get("exit_date"),
                t.get("exit_price"),
                t.get("shares"),
                t.get("cost"),
                t.get("realized_pnl"),
                t.get("return_pct"),
                holding_days_calendar(t.get("entry_date"), t.get("exit_date")),
                normalize_exit_reason(t.get("exit_reason")) or t.get("exit_reason"),
                t.get("stop_price"),
                t.get("take_profit_price"),
                t.get("ai_score_entry"),
                fin,
                t.get("news_entry") or t.get("news_tone_entry"),
                t.get("knife_score_entry") or t.get("knife_score"),
                t.get("range_63d_pos_entry"),
                "Y" if t.get("is_priority") else "",
            ]
        )
    _write_sheet(wb, "Trade History", hist_headers, hist_rows)

    open_headers = [
        "Ticker",
        "Source",
        "Entry Date",
        "Entry Price",
        "Shares",
        "Cost",
        "Current Price",
        "Market Value",
        "Unrealized P&L",
        "Unrealized %",
        "Stop Loss",
        "Take Profit",
        "AI Score at Entry",
        "Current AI Score",
        "Priority Buy",
    ]
    open_rows = [
        [
            t.get("ticker"),
            format_source_label(t.get("source_at_entry")),
            t.get("entry_date"),
            t.get("entry_price"),
            t.get("shares"),
            t.get("cost"),
            t.get("current_price"),
            t.get("market_value"),
            t.get("unrealized_pnl"),
            t.get("unrealized_pnl_pct"),
            t.get("stop_price"),
            t.get("take_profit_price"),
            t.get("ai_score_entry"),
            t.get("ai_score_current"),
            "Y" if t.get("is_priority") else "",
        ]
        for t in opens
    ]
    _write_sheet(wb, "Open Positions", open_headers, open_rows)

    disc_headers = [
        "Ticker",
        "Event",
        "Source",
        "Event Score",
        "Direction",
        "Event Date",
        "Period",
        "Status",
        "Discovery Price",
        "AI Score",
        "Knife",
        "PRIORITY (Saved)",
        "Recent",
    ]
    disc_rows = [
        [
            r.get("ticker"),
            r.get("event_summary"),
            r.get("source_display") or r.get("primary_source") or r.get("source_tags"),
            r.get("event_score"),
            r.get("sentiment"),
            r.get("event_date"),
            r.get("event_period"),
            r.get("status"),
            r.get("discovery_price") or r.get("price"),
            r.get("ai_score"),
            r.get("knife_score"),
            "Y" if r.get("is_news_priority") else "",
            "Y" if r.get("is_recent") else "",
        ]
        for r in discovery
    ]
    _write_sheet(wb, "AI Discovery", disc_headers, disc_rows)

    saved_headers = [
        "Ticker",
        "Event",
        "Source",
        "Event Score",
        "Direction",
        "Event Date",
        "Period",
        "Status",
        "AI Score",
        "Knife",
        "Saved At",
    ]
    saved_rows = [
        [
            r.get("ticker"),
            r.get("event_summary"),
            r.get("source_display") or r.get("primary_source") or r.get("source_tags"),
            r.get("event_score"),
            r.get("sentiment"),
            r.get("event_date"),
            r.get("event_period"),
            r.get("status"),
            r.get("ai_score"),
            r.get("knife_score"),
            r.get("news_priority_at"),
        ]
        for r in saved
    ]
    _write_sheet(wb, "Saved News", saved_headers, saved_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def archive_legacy_ai_trading() -> dict[str, Any]:
    """
    Export active paper-trading experiment tables to
    data/logs/LEGACY_AI_TRADING_<stamp>.json before reset.

    Never touches daily_bars / universe / Strong / Rising / Sector Rotation / Watchlist.
    """
    import json
    from pathlib import Path

    init_db()
    now = _utc_now_iso()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    root = Path(__file__).resolve().parent / "data" / "logs"
    root.mkdir(parents=True, exist_ok=True)
    path = root / f"LEGACY_AI_TRADING_{stamp}.json"
    tables = (
        "paper_trades",
        "paper_candidates",
        "paper_priority",
        "paper_equity_snapshots",
        "paper_level_overrides",
        "paper_portfolio",
    )
    payload: dict[str, Any] = {
        "archive_batch": stamp,
        "archived_at": now,
        "label": "LEGACY_AI_TRADING",
        "tables": {},
    }
    counts: dict[str, int] = {}
    with get_conn() as conn:
        for table in tables:
            try:
                rows = conn.execute(f"SELECT * FROM {table}").fetchall()
            except Exception:
                counts[table] = 0
                continue
            payload["tables"][table] = [dict(r) for r in rows]
            counts[table] = len(payload["tables"][table])
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return {
        "archive_batch": stamp,
        "archived_at": now,
        "path": str(path),
        "counts": counts,
    }


def reset_ai_trading(*, archive_first: bool = True) -> dict[str, Any]:
    """
    Wipe the AI Paper Trading experiment and restore cash to starting capital.

    With archive_first=True (default for architecture migration), copies rows into
    LEGACY_* tables before delete.

    Does NOT delete: Watchlist membership, Research history, daily_bars, universe,
    Strong/Rising/Sector Rotation, AI Discovery research events (unlink paper ids only),
    financial/news caches, system settings other than paper_* as_of keys.
    """
    from paper_trading import ensure_portfolio, _cfg

    archive_info: dict[str, Any] = {}
    if archive_first:
        try:
            archive_info = archive_legacy_ai_trading()
        except Exception as exc:
            log.exception("LEGACY archive failed")
            archive_info = {"ok": False, "error": str(exc)}

    init_db()
    cfg = _cfg()
    start = float(cfg["starting_capital"])
    now = _utc_now_iso()

    with get_conn() as conn:
        n_trades = conn.execute("SELECT COUNT(*) AS n FROM paper_trades").fetchone()["n"]
        n_open = conn.execute(
            "SELECT COUNT(*) AS n FROM paper_trades WHERE status = 'open'"
        ).fetchone()["n"]
        n_closed = conn.execute(
            "SELECT COUNT(*) AS n FROM paper_trades WHERE status = 'closed'"
        ).fetchone()["n"]
        n_pri = conn.execute("SELECT COUNT(*) AS n FROM paper_priority").fetchone()["n"]
        n_cand = conn.execute("SELECT COUNT(*) AS n FROM paper_candidates").fetchone()["n"]
        n_snap = conn.execute(
            "SELECT COUNT(*) AS n FROM paper_equity_snapshots"
        ).fetchone()["n"]

        # Unlink Discovery rows from trades we are about to delete (keep research).
        conn.execute(
            """
            UPDATE ai_discovery_candidates
            SET paper_trade_id = NULL,
                status = CASE
                  WHEN status = 'ORDER_CREATED' THEN 'WATCH'
                  ELSE status
                END,
                updated_at = ?
            WHERE paper_trade_id IS NOT NULL
            """,
            (now,),
        )

        conn.execute("DELETE FROM paper_level_overrides")
        conn.execute("DELETE FROM paper_trades")
        conn.execute("DELETE FROM paper_priority")
        conn.execute("DELETE FROM paper_candidates")
        conn.execute("DELETE FROM paper_equity_snapshots")
        # Clear temporary new-architecture snapshots for a clean active environment.
        # Permanent market / Strong / Rising / Sector Rotation / Watchlist / Discovery kept.
        try:
            conn.execute("DELETE FROM ai_buy_snapshots")
        except Exception:
            pass
        try:
            conn.execute("DELETE FROM ai_select_candidates")
        except Exception:
            pass
        try:
            conn.execute("DELETE FROM trading_order_requests")
        except Exception:
            pass
        conn.execute(
            """
            UPDATE paper_portfolio
            SET cash = ?, starting_capital = ?, trading_limit = ?,
                reserve_cash = ?, updated_at = ?
            WHERE id = 1
            """,
            (
                start,
                start,
                float(cfg["trading_limit"]),
                float(cfg["reserve_cash"]),
                now,
            ),
        )

    # Ensure portfolio row exists after wipe.
    ensure_portfolio()
    set_keys = (
        "paper_candidates_as_of",
        "paper_last_daily_update",
        "ai_select_as_of",
        "ai_select_built_at",
        "ai_buy_as_of",
        "ai_buy_built_at",
    )
    from db import set_setting

    for k in set_keys:
        try:
            set_setting(k, "")
        except Exception:
            pass

    return {
        "trades_deleted": int(n_trades or 0),
        "open_deleted": int(n_open or 0),
        "closed_deleted": int(n_closed or 0),
        "priority_cleared": int(n_pri or 0),
        "candidates_cleared": int(n_cand or 0),
        "snapshots_cleared": int(n_snap or 0),
        "cash_restored": start,
        "reset_at": now,
        "legacy_archive": archive_info,
        "architecture": "AI_SELECT_AI_BUY",
    }
